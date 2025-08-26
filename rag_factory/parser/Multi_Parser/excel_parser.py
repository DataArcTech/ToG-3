import logging
import re
import sys
from io import BytesIO
import os
import pandas as pd
import chardet
from openpyxl import Workbook, load_workbook

class ExcelParser:

    def _dataframe_to_workbook(self, df):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        for col_num, column_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)

        for row_num, row in enumerate(df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        return wb

    def _load_excel_to_workbook(self, file_like_object):
        if isinstance(file_like_object, bytes):
            file_like_object = BytesIO(file_like_object)
        elif isinstance(file_like_object, str):
            if not os.path.exists(file_like_object):
                raise FileNotFoundError(f"Excel 文件不存在: {file_like_object}")
            file_like_object = open(file_like_object, "rb")

        # Read first 4 bytes to determine file type
        file_like_object.seek(0)
        file_head = file_like_object.read(4)
        file_like_object.seek(0)

        if not (file_head.startswith(b'PK\x03\x04') or file_head.startswith(b'\xD0\xCF\x11\xE0')):
            logging.info("Converting CSV to Excel Workbook")
            file_like_object.seek(0)
            raw = file_like_object.read(1000) 
            result = chardet.detect(raw)
            encoding = result['encoding'] or "utf-8"
            file_like_object.seek(0)
            try:
                file_like_object.seek(0)
                df = pd.read_csv(file_like_object, encoding=encoding)
                return self._dataframe_to_workbook(df)

            except Exception as e:
                raise Exception(f"Failed to parse CSV and convert to Excel Workbook: {e}")

        try:
            return load_workbook(file_like_object,data_only= True)
        except Exception as e:
            logging.info(f"openpyxl load error: {e}, try pandas instead")
            try:
                file_like_object.seek(0)
                try:
                    df = pd.read_excel(file_like_object)
                    return self._dataframe_to_workbook(df)
                except Exception as ex:
                    logging.info(f"pandas with default engine load error: {ex}, try calamine instead")
                    file_like_object.seek(0)
                    df = pd.read_excel(file_like_object, engine='calamine')
                    return self._dataframe_to_workbook(df)
            except Exception as e_pandas:
                raise Exception(f"pandas.read_excel error: {e_pandas}, original openpyxl error: {e}")


    def _html(self, fnm, chunk_rows=256):
        from html import escape

        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = self._load_excel_to_workbook(file_like_object)
        tb_chunks = []

        def _fmt(v):
            if v is None:
                return ""
            return str(v).strip()

        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue

            tb_rows_0 = "<tr>"
            for t in list(rows[0]):
                tb_rows_0 += f"<th>{escape(_fmt(t.value))}</th>"
            tb_rows_0 += "</tr>"

            for chunk_i in range((len(rows) - 1) // chunk_rows + 1):
                tb = ""
                tb += f"<table><caption>{sheetname}</caption>"
                tb += tb_rows_0
                for r in list(
                    rows[1 + chunk_i * chunk_rows: min(1 + (chunk_i + 1) * chunk_rows, len(rows))]
                ):
                    tb += "<tr>"
                    for i, c in enumerate(r):
                        if c.value is None:
                            tb += "<td></td>"
                        else:
                            tb += f"<td>{c.value}</td>"
                    tb += "</tr>"
                tb += "</table>\n"
                tb_chunks.append(tb)

        return tb_chunks
    
    def __call__(self, fnm, output_dir="./output"):
        md_output = self._html(fnm)
        if isinstance(fnm, str):
            file_name = os.path.basename(fnm).rsplit(".", 1)[0]
        elif isinstance(fnm, bytes):
            prefix_bytes = fnm[:10]
            prefix_hex = prefix_bytes.hex()
            # prefix_b64 = base64.urlsafe_b64encode(prefix_bytes).decode('utf-8')
            file_name= f"{prefix_hex}"
        else:
            file_name= "input_excel" 
        output_path = os.path.join(output_dir, file_name+"_markdown.md")
        
        with open(output_path, 'w', encoding='utf-8') as output_f:
            for item in md_output:
                output_f.write(item)
                output_f.write("/n/n")
                

if __name__ == "__main__":
    file_name = '/home/yangcehao/doc_analysis/MultiTypeParser/地址解析验证.xlsx'
    parser = ExcelParser()
    parser(file_name, output_dir='/home/yangcehao/doc_analysis/MultiTypeParser/test')
